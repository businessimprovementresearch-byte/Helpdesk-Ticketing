from django.contrib.auth.decorators import login_required
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.management import call_command
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Avg, Count, F, ExpressionWrapper, DurationField
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from django.conf import settings
import io
import contextlib
from datetime import timedelta

@login_required
def dashboard(request):
    visible = _visible_tickets_for(request.user)

    closed_qs = visible.filter(status=Ticket.Status.CLOSED).annotate(
        resolution_time=ExpressionWrapper(F("updated_at") - F("created_at"), output_field=DurationField())
    )
    avg_duration = closed_qs.aggregate(avg=Avg("resolution_time"))["avg"]
    avg_hours = round(avg_duration.total_seconds() / 3600, 1) if avg_duration else 0

    return render(request, "tickets/dashboard.html", {
        "open_count": visible.filter(status=Ticket.Status.OPEN).count(),
        "pending_count": visible.filter(status=Ticket.Status.PENDING).count(),
        "closed_count": visible.filter(status=Ticket.Status.CLOSED).count(),
        "avg_hours": avg_hours,
        "recent_tickets": visible.select_related("created_by")[:5],
    })

from accounts.models import Division
from .emails import (
    send_new_ticket_notification,
    send_reply_notification,
    send_status_change_notification,
)
from .forms import (
    AttachmentUploadForm,
    DivisionForm,
    TicketForm,
    TicketReplyForm,
    TicketStatusForm,
    UserForm,
    validate_attachment_files,
)
from .models import Attachment, Ticket, TicketReply

User = get_user_model()


def _admin_required(user):
    if not user.is_admin:
        raise PermissionDenied("Halaman ini khusus admin.")


def _visible_tickets_for(user):
    """Customer cuma lihat tiket miliknya sendiri; agent/admin lihat semua."""
    if user.is_customer:
        return Ticket.objects.filter(created_by=user)
    return Ticket.objects.all()

def _filter_tickets_by_period(tickets, request):
    period = request.GET.get("period", "all")
    now = timezone.now()
    if period == "today":
        return tickets.filter(created_at__date=now.date()), "HARI INI"
    if period == "week":
        return tickets.filter(created_at__gte=now - timedelta(days=7)), "MINGGU INI"
    if period == "month":
        return tickets.filter(created_at__year=now.year, created_at__month=now.month), "BULAN INI"
    if period == "custom":
        date_from = request.GET.get("date_from")
        date_to = request.GET.get("date_to")
        qs = tickets
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        label = f"{date_from or '...'} s/d {date_to or '...'}"
        return qs, label
    return tickets, "SEMUA"


@login_required
def index(request):
    tickets = _visible_tickets_for(request.user).select_related("division", "assigned_to")

    status = request.GET.get("status")
    priority = request.GET.get("priority")
    q = request.GET.get("q")
    division = request.GET.get("division")

    if status:
        tickets = tickets.filter(status=status)
    if priority:
        tickets = tickets.filter(priority=priority)
    if q:
        tickets = tickets.filter(subject__icontains=q)
    if division:
        tickets = tickets.filter(division_id=division)

    paginator = Paginator(tickets, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "tickets/index.html", {
        "page_obj": page_obj,
        "status_options": [{"value": v, "label": l, "selected": v == (status or "")} for v, l in Ticket.Status.choices],
        "priority_options": [{"value": v, "label": l, "selected": v == (priority or "")} for v, l in Ticket.Priority.choices],
        "current_status": status or "",
        "current_priority": priority or "",
        "q": q or "",
    })


@login_required
def ticket_create(request):
    if request.method == "POST":
        form = TicketForm(request.POST, user=request.user)
        attachment_form = AttachmentUploadForm(request.POST, request.FILES)
        uploaded_files = request.FILES.getlist("files")

        attachment_error = None
        if attachment_form.is_valid():
            try:
                validate_attachment_files(uploaded_files)
            except ValidationError as exc:
                attachment_error = exc.message

        if form.is_valid() and attachment_form.is_valid() and not attachment_error:
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            ticket.source = Ticket.Source.WEB
            # Ticket code digenerate otomatis berdasarkan divisi yang dipilih
            ticket.ticket_code = Ticket.generate_ticket_code(division=ticket.division)
            ticket.save()

            reply = TicketReply.objects.create(
                ticket=ticket,
                author=request.user,
                type=TicketReply.Type.PUBLIC_REPLY,
                message=form.cleaned_data["message"],
            )

            for uploaded_file in uploaded_files:
                Attachment.objects.create(
                    reply=reply,
                    file=uploaded_file,
                    file_name=uploaded_file.name,
                    mime_type=getattr(uploaded_file, "content_type", "") or "",
                    file_size=uploaded_file.size,
                )

            send_new_ticket_notification(ticket)
            messages.success(request, f"Tiket {ticket.ticket_code} berhasil dibuat.")
            return redirect("tickets:detail", pk=ticket.pk)
        elif attachment_error:
            messages.error(request, attachment_error)
    else:
        form = TicketForm(user=request.user)
        attachment_form = AttachmentUploadForm()

    return render(request, "tickets/ticket_form.html", {
        "form": form,
        "attachment_form": attachment_form,
    })


@login_required
def ticket_detail(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)

    if request.user.is_customer and ticket.created_by_id != request.user.id:
        raise PermissionDenied("Kamu tidak punya akses ke tiket ini.")

    can_use_internal_note = not request.user.is_customer
    can_manage_status = not request.user.is_customer  # agent & admin

    status_form = TicketStatusForm(instance=ticket) if can_manage_status else None

    if request.method == "POST":
        action = request.POST.get("action", "reply")

        # --- Aksi 1: agent/admin mengubah status/priority/assignment tiket ---
        if action == "update_status" and can_manage_status:
            old_status = ticket.status
            status_form = TicketStatusForm(request.POST, instance=ticket)
            if status_form.is_valid():
                updated_ticket = status_form.save()
                send_status_change_notification(updated_ticket, old_status, changed_by=request.user)
                messages.success(request, "Status tiket diperbarui.")
                return redirect("tickets:detail", pk=ticket.pk)
            else:
                messages.error(request, "Gagal memperbarui status tiket.")

        # --- Aksi 2: kirim balasan (customer atau agent/admin) ---
        else:
            reply_form = TicketReplyForm(request.POST, can_use_internal_note=can_use_internal_note)
            attachment_form = AttachmentUploadForm(request.POST, request.FILES)
            uploaded_files = request.FILES.getlist("files")

            attachment_error = None
            try:
                validate_attachment_files(uploaded_files)
            except ValidationError as exc:
                attachment_error = exc.message

            if reply_form.is_valid() and attachment_form.is_valid() and not attachment_error:
                reply = reply_form.save(commit=False)
                reply.ticket = ticket
                reply.author = request.user
                if not can_use_internal_note:
                    reply.type = TicketReply.Type.PUBLIC_REPLY
                reply.save()

                for uploaded_file in uploaded_files:
                    Attachment.objects.create(
                        reply=reply,
                        file=uploaded_file,
                        file_name=uploaded_file.name,
                        mime_type=getattr(uploaded_file, "content_type", "") or "",
                        file_size=uploaded_file.size,
                    )

                # Auto-reopen: kalau tiket closed dan yang membalas adalah
                # customer pemilik tiket (bukan internal note), buka lagi otomatis.
                if (
                    ticket.status == Ticket.Status.CLOSED
                    and reply.type == TicketReply.Type.PUBLIC_REPLY
                    and request.user.id == ticket.created_by_id
                ):
                    old_status = ticket.status
                    ticket.status = Ticket.Status.OPEN
                    ticket.save(update_fields=["status", "updated_at"])
                    send_status_change_notification(ticket, old_status, changed_by=request.user)
                    messages.info(request, "Tiket dibuka kembali karena ada balasan baru.")

                send_reply_notification(reply)
                messages.success(request, "Balasan terkirim.")
                return redirect("tickets:detail", pk=ticket.pk)
            elif attachment_error:
                messages.error(request, attachment_error)
    else:
        reply_form = TicketReplyForm(can_use_internal_note=can_use_internal_note)
        attachment_form = AttachmentUploadForm()

    replies = ticket.replies.select_related("author").prefetch_related("attachments")
    if request.user.is_customer:
        replies = replies.exclude(type=TicketReply.Type.INTERNAL_NOTE)

    return render(request, "tickets/ticket_detail.html", {
        "ticket": ticket,
        "replies": replies,
        "reply_form": reply_form,
        "attachment_form": attachment_form,
        "status_form": status_form,
        "can_manage_status": can_manage_status,
    })


@login_required
def attachment_download(request, pk):
    attachment = get_object_or_404(Attachment, pk=pk)
    ticket = attachment.reply.ticket

    if request.user.is_customer and ticket.created_by_id != request.user.id:
        raise PermissionDenied("Kamu tidak punya akses ke file ini.")
    if request.user.is_customer and attachment.reply.type == TicketReply.Type.INTERNAL_NOTE:
        raise Http404("File tidak ditemukan.")

    # Kalau pakai remote storage (Supabase S3), redirect ke signed URL-nya.
    # Kalau local storage, stream langsung filenya.
    if hasattr(attachment.file, "url") and attachment.file.storage.__class__.__name__ != "FileSystemStorage":
        return redirect(attachment.file.url)

    return FileResponse(attachment.file.open("rb"), as_attachment=True, filename=attachment.file_name)


# ---------------------------------------------------------------------------
# Users (full CRUD + bulk delete) — khusus admin
# ---------------------------------------------------------------------------

@login_required
def users_index(request):
    _admin_required(request.user)

    if request.method == "POST" and request.POST.get("action") == "change_role":
        target = get_object_or_404(User, pk=request.POST.get("user_id"))
        new_role = request.POST.get("new_role")
        if new_role in dict(User.Role.choices):
            target.role = new_role
            target.save(update_fields=["role"])
            messages.success(request, "Peran pengguna berhasil diubah.")
        return redirect("tickets:users")

    if request.method == "POST" and request.POST.get("action") == "bulk_delete":
        selected_ids = request.POST.getlist("selected_users")
        # jaga-jaga: jangan pernah izinkan user menghapus akunnya sendiri lewat bulk action
        selected_ids = [pk for pk in selected_ids if str(pk) != str(request.user.pk)]
        if selected_ids:
            deleted_count, _ = User.objects.filter(pk__in=selected_ids).delete()
            messages.success(request, f"{len(selected_ids)} pengguna berhasil dihapus.")
        else:
            messages.warning(request, "Tidak ada pengguna yang dipilih untuk dihapus.")
        return redirect("tickets:users")

    q = request.GET.get("q")
    role = request.GET.get("role")

    users = User.objects.all().select_related("division").order_by("first_name", "username")
    if q:
        users = users.filter(Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(email__icontains=q) | Q(username__icontains=q))
    if role:
        users = users.filter(role=role)

    paginator = Paginator(users, 10)
    page_obj = paginator.get_page(request.GET.get("page"))

    return render(request, "tickets/users_index.html", {
        "page_obj": page_obj,
        "q": q or "",
        "current_role": role or "",
        "role_options": [
            {"value": v, "label": l, "selected": v == (role or "")}
            for v, l in [("customer", "User"), ("agent", "Operator"), ("admin", "Admin")]
        ],
        "create_form": UserForm(),
    })


@login_required
def user_create(request):
    _admin_required(request.user)
    if request.method == "POST":
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "User baru berhasil dibuat.")
            return redirect("tickets:users")
    else:
        form = UserForm()
    return render(request, "tickets/user_form.html", {"form": form, "is_new": True})


@login_required
def user_edit(request, pk):
    _admin_required(request.user)
    target_user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        form = UserForm(request.POST, instance=target_user)
        if form.is_valid():
            form.save()
            messages.success(request, "User berhasil diperbarui.")
            return redirect("tickets:users")
    else:
        form = UserForm(instance=target_user)
    return render(request, "tickets/user_form.html", {"form": form, "is_new": False, "target_user": target_user})


@login_required
def user_delete(request, pk):
    _admin_required(request.user)
    target_user = get_object_or_404(User, pk=pk)
    if request.method == "POST":
        if target_user.pk == request.user.pk:
            messages.error(request, "Kamu tidak bisa menghapus akunmu sendiri.")
        else:
            target_user.delete()
            messages.success(request, "User berhasil dihapus.")
    return redirect("tickets:users")


# ---------------------------------------------------------------------------
# Divisions (full CRUD)
# ---------------------------------------------------------------------------

@login_required
def division_create(request):
    _admin_required(request.user)
    if request.method == "POST":
        form = DivisionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Divisi baru berhasil dibuat.")
            return redirect("tickets:divisions")
    else:
        form = DivisionForm()
    return render(request, "tickets/division_form.html", {"form": form, "is_new": True})


@login_required
def division_edit(request, pk):
    _admin_required(request.user)
    division = get_object_or_404(Division, pk=pk)
    if request.method == "POST":
        form = DivisionForm(request.POST, instance=division)
        if form.is_valid():
            form.save()
            messages.success(request, "Divisi berhasil diperbarui.")
            return redirect("tickets:divisions")
    else:
        form = DivisionForm(instance=division)
    return render(request, "tickets/division_form.html", {"form": form, "is_new": False, "division": division})


@login_required
def division_toggle_active(request, pk):
    _admin_required(request.user)
    division = get_object_or_404(Division, pk=pk)
    if request.method == "POST":
        division.is_active = not division.is_active
        division.save(update_fields=["is_active"])
        messages.success(
            request,
            f"Divisi {division.name} sekarang {'aktif' if division.is_active else 'nonaktif'}.",
        )
    return redirect("tickets:divisions")

@login_required
def division_delete(request, pk):
    _admin_required(request.user)
    division = get_object_or_404(Division, pk=pk)
    if request.method == "POST":
        name = division.name
        division.delete()
        messages.success(request, f"Divisi {name} berhasil dihapus.")
    return redirect("tickets:divisions")

@login_required
def divisions_index(request):
    open_form = False
    if request.method == "POST":
        _admin_required(request.user)
        form = DivisionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Divisi baru berhasil dibuat.")
            return redirect("tickets:divisions")
        open_form = True
    else:
        form = DivisionForm()

    q = request.GET.get("q", "").strip()
    divisions = Division.objects.annotate(ticket_count=Count("tickets")).order_by("name")
    if q:
        divisions = divisions.filter(Q(name__icontains=q) | Q(code__icontains=q))

    return render(request, "tickets/divisions_index.html", {
        "divisions": divisions,
        "form": form,
        "q": q,
        "open_form": open_form,
    })

# ---------------------------------------------------------------------------
# Reports (+ export Excel)
# ---------------------------------------------------------------------------

@login_required
def reports_index(request):
    tickets = _visible_tickets_for(request.user)
    now = timezone.now()

    tickets_period, _ = _filter_tickets_by_period(tickets, request)

    summary = {
        "total": tickets_period.count(),
        "open": tickets_period.filter(status=Ticket.Status.OPEN).count(),
        "pending": tickets_period.filter(status=Ticket.Status.PENDING).count(),
        "closed": tickets_period.filter(status=Ticket.Status.CLOSED).count(),
        "needs_attention": tickets.filter(priority=Ticket.Priority.HIGH).exclude(status=Ticket.Status.CLOSED).count(),
        "this_month": tickets.filter(created_at__year=now.year, created_at__month=now.month).count(),
    }

    return render(request, "tickets/reports_index.html", {
        "summary": summary,
        "current_period": request.GET.get("period", "all"),
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
        "recent_tickets": tickets.filter(created_at__gte=now - timedelta(days=7)).select_related("created_by").order_by("-created_at"),
    })

@login_required
def reports_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    tickets = _visible_tickets_for(request.user).select_related("division", "created_by")
    tickets, period_label = _filter_tickets_by_period(tickets, request)
    tickets = tickets.order_by("-created_at")

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Tiket"

    headers = ["Nama Perusahaan", "Customer Name", "Issue", "Salesman", "Invoice (Y/N)", "Status"]
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value=f"LAPORAN TIKET {period_label} - {timezone.now().strftime('%d %B %Y').upper()}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.fill = PatternFill("solid", fgColor="8DB4E2")

    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=text)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")
        cell.alignment = Alignment(horizontal="center")

    row_idx = 3
    for t in tickets:
        invoice_raw = (t.invoice_status or "").strip().lower()
        invoice_yn = "Y" if "sudah" in invoice_raw or invoice_raw in ("y", "yes", "lunas") else ("N" if invoice_raw else "")
        ws.append([
            t.nama_perusahaan or (str(t.division) if t.division else "-"),
            t.created_by.get_full_name() if t.created_by else "-",
            t.subject,
            t.salesman,
            invoice_yn,
            t.get_status_display(),
        ])
        row_idx += 1

    widths = {"A": 38, "B": 20, "C": 63, "D": 13, "E": 18, "F": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    from django.http import HttpResponse
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"laporan-tiket-{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ---------------------------------------------------------------------------
# Webhook untuk trigger fetch_emails dari luar (GitHub Actions scheduled
# workflow, gratis) — pengganti Render Cron Job yang berbayar.
# Dilindungi oleh CRON_SECRET, bukan login, karena yang manggil bukan user.
# ---------------------------------------------------------------------------

@csrf_exempt
@require_GET
def fetch_emails_webhook(request):
    token = request.GET.get("token", "")
    expected = getattr(settings, "CRON_SECRET", "")
    if not expected or token != expected:
        raise Http404()

    buffer = io.StringIO()
    with contextlib.redirect_stdout(buffer), contextlib.redirect_stderr(buffer):
        call_command("fetch_emails")

    return JsonResponse({"ok": True, "log": buffer.getvalue()})